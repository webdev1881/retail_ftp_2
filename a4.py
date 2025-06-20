#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FTP Analyzer - Этап 4: Сравнительный отчет двух периодов
Автор: Система аналитики данных
Версия: 4.0
"""

import ftplib
import os
import json
import sys
import csv
import io
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd

class FTPAnalyzer:
    def __init__(self):
        # Настройки FTP-сервера
        self.ftp_host = "smkft.space"
        self.ftp_port = 2122
        self.ftp_user = "nielsen"
        self.ftp_pass = "Qazwsx32123"
        
        # Настройки городов
        self.cities = {
            'khar': 'Харьков',
            'kiev': 'Киев', 
            'dnepr': 'Днепр',
            'bel': 'Белая Церковь'
        }
        
        # Настройки локальных папок
        self.base_dir = Path("ftp_data")
        self.cache_dir = self.base_dir / "cache"
        self.reports_dir = self.base_dir / "reports"
        self.structure_file = self.base_dir / "ftp_structure.json"
        
        # Создаем необходимые папки
        self.create_directories()
        
    def create_directories(self):
        """Создает необходимые папки для работы программы"""
        try:
            self.base_dir.mkdir(exist_ok=True)
            self.cache_dir.mkdir(exist_ok=True)
            self.reports_dir.mkdir(exist_ok=True)
            print(f"✓ Создана рабочая папка: {self.base_dir.absolute()}")
            print(f"✓ Создана папка кеша: {self.cache_dir.absolute()}")
            print(f"✓ Создана папка отчетов: {self.reports_dir.absolute()}")
        except Exception as e:
            print(f"✗ Ошибка создания папок: {e}")
            sys.exit(1)
            
    def clear_cache(self):
        """Очищает кеш"""
        try:
            import shutil
            if self.cache_dir.exists():
                shutil.rmtree(self.cache_dir)
                self.cache_dir.mkdir()
                print("✓ Кеш очищен")
                return True
        except Exception as e:
            print(f"✗ Ошибка очистки кеша: {e}")
            return False
            
    def ask_clear_cache(self):
        """Спрашивает пользователя об очистке кеша"""
        while True:
            choice = input("\n🗑️  Очистить кеш перед загрузкой? (y/n): ").lower().strip()
            if choice in ['y', 'yes', 'д', 'да']:
                return self.clear_cache()
            elif choice in ['n', 'no', 'н', 'нет']:
                print("ℹ️  Кеш сохранен")
                return True
            else:
                print("❌ Введите 'y' для да или 'n' для нет")
                
    def connect_ftp(self):
        """Подключается к FTP-серверу"""
        try:
            print(f"🔄 Подключаемся к FTP-серверу {self.ftp_host}:{self.ftp_port}...")
            ftp = ftplib.FTP()
            ftp.connect(self.ftp_host, self.ftp_port)
            ftp.login(self.ftp_user, self.ftp_pass)
            print("✓ Подключение к FTP установлено")
            return ftp
        except ftplib.error_perm as e:
            print(f"✗ Ошибка авторизации FTP: {e}")
            return None
        except Exception as e:
            print(f"✗ Ошибка подключения к FTP: {e}")
            return None
            
    def download_file(self, ftp, remote_path, local_path):
        """Скачивает файл с FTP сервера"""
        try:
            # Проверяем, есть ли файл в кеше
            if local_path.exists():
                print(f"📄 Файл найден в кеше: {local_path.name}")
                return True
                
            print(f"⬇️  Скачиваем: {remote_path}")
            local_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(local_path, 'wb') as f:
                ftp.retrbinary(f'RETR {remote_path}', f.write)
            
            print(f"✓ Скачан: {local_path.name}")
            return True
            
        except ftplib.error_perm as e:
            if "550" in str(e):  # Файл не найден
                print(f"⚠️  Файл не найден: {remote_path}")
            else:
                print(f"✗ Ошибка FTP при скачивании {remote_path}: {e}")
            return False
        except Exception as e:
            print(f"✗ Ошибка скачивания {remote_path}: {e}")
            return False
            
    def read_csv_with_pipe(self, file_path):
        """Читает CSV файл с разделителем |"""
        try:
            df = pd.read_csv(file_path, sep='|', encoding='utf-8')
            
            # Преобразуем числовые поля в правильные типы
            file_name = file_path.name.lower()
            
            if 'cartitem' in file_name:
                # Для файлов cartitem преобразуем qty и total_price в числа
                if 'qty' in df.columns:
                    # Обрабатываем числа с запятой как разделителем десятичных
                    df['qty'] = df['qty'].astype(str).str.replace(',', '.', regex=False)
                    df['qty'] = pd.to_numeric(df['qty'], errors='coerce').fillna(0)
                if 'total_price' in df.columns:
                    # Обрабатываем числа с запятой как разделителем десятичных
                    df['total_price'] = df['total_price'].astype(str).str.replace(',', '.', regex=False)
                    df['total_price'] = pd.to_numeric(df['total_price'], errors='coerce').fillna(0.0)
            
            # Убираем пробелы из строковых полей
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.strip()
            
            print(f"✓ Прочитан CSV: {file_path.name} ({len(df)} строк)")
            return df
        except Exception as e:
            print(f"✗ Ошибка чтения CSV {file_path}: {e}")
            return None
            
    def generate_date_range(self, start_date, end_date):
        """Генерирует список дат в указанном диапазоне"""
        dates = []
        current_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        
        while current_date <= end_date_obj:
            dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
            
        return dates
        
    def download_shop_files(self, ftp):
        """Скачивает справочники магазинов"""
        print("\n📋 ЗАГРУЗКА СПРАВОЧНИКОВ МАГАЗИНОВ")
        print("-" * 40)
        
        shops_data = {}
        
        for city_code, city_name in self.cities.items():
            remote_path = f"/www/shop_{city_code}.csv"
            local_path = self.cache_dir / f"shop_{city_code}.csv"
            
            if self.download_file(ftp, remote_path, local_path):
                df = self.read_csv_with_pipe(local_path)
                if df is not None:
                    shops_data[city_code] = df
                    print(f"  📊 {city_name}: {len(df)} магазинов")
                    
        return shops_data
        
    def download_period_data(self, ftp, start_date, end_date, period_name):
        """Скачивает данные за указанный период"""
        print(f"\n📅 ЗАГРУЗКА ДАННЫХ ЗА {period_name.upper()} ({start_date} - {end_date})")
        print("-" * 70)
        
        dates = self.generate_date_range(start_date, end_date)
        receipts_data = {}
        cartitems_data = {}
        
        for city_code, city_name in self.cities.items():
            print(f"\n🏙️  {city_name} ({city_code}):")
            receipts_data[city_code] = []
            cartitems_data[city_code] = []
            
            for date in dates:
                # Скачиваем чеки
                receipt_remote = f"/www/receipt/receipt_{city_code}_{date}.csv"
                receipt_local = self.cache_dir / f"receipt_{city_code}_{date}.csv"
                
                # Скачиваем товарные позиции
                cartitem_remote = f"/www/cartitem/cartitem_{city_code}_{date}.csv"
                cartitem_local = self.cache_dir / f"cartitem_{city_code}_{date}.csv"
                
                receipt_loaded = False
                cartitem_loaded = False
                
                # Загружаем чеки
                if self.download_file(ftp, receipt_remote, receipt_local):
                    df = self.read_csv_with_pipe(receipt_local)
                    if df is not None:
                        df['date'] = date
                        receipts_data[city_code].append(df)
                        receipt_loaded = True
                        
                # Загружаем товарные позиции
                if self.download_file(ftp, cartitem_remote, cartitem_local):
                    df = self.read_csv_with_pipe(cartitem_local)
                    if df is not None:
                        df['date'] = date
                        cartitems_data[city_code].append(df)
                        cartitem_loaded = True
                        
                if receipt_loaded and cartitem_loaded:
                    receipt_df = receipts_data[city_code][-1]
                    cartitem_df = cartitems_data[city_code][-1]
                    total_qty = float(cartitem_df['qty'].sum()) if 'qty' in cartitem_df.columns else 0.0
                    total_price = float(cartitem_df['total_price'].sum()) if 'total_price' in cartitem_df.columns else 0.0
                    print(f"    📅 {date}: {len(receipt_df)} чеков, {len(cartitem_df)} позиций, товаров: {total_qty:.1f}, оборот: {total_price:.2f}")
                    
        return receipts_data, cartitems_data
        
    def calculate_period_metrics(self, shops_data, receipts_data, cartitems_data, period_name):
        """Рассчитывает метрики для одного периода"""
        print(f"\n📊 РАСЧЕТ МЕТРИК ДЛЯ {period_name.upper()}")
        print("-" * 50)
        
        period_metrics = {}
        
        for city_code, city_name in self.cities.items():
            print(f"\n🏙️  Обрабатываем {city_name}...")
            
            if city_code not in shops_data:
                print(f"  ⚠️  Нет данных о магазинах для {city_name}")
                continue
                
            shops_df = shops_data[city_code]
            
            # Инициализируем метрики нулями
            city_metrics = {}
            for _, shop in shops_df.iterrows():
                city_metrics[shop['id']] = {
                    'shop_name': shop['name'],
                    'receipts': 0,
                    'revenue': 0.0,
                    'quantity': 0.0
                }
            
            # Обрабатываем чеки
            if city_code in receipts_data and receipts_data[city_code]:
                all_receipts = pd.concat(receipts_data[city_code], ignore_index=True)
                receipt_counts = all_receipts.groupby('shop_id').size().reset_index(name='receipt_count')
                
                for _, row in receipt_counts.iterrows():
                    shop_id = row['shop_id']
                    if shop_id in city_metrics:
                        city_metrics[shop_id]['receipts'] = int(row['receipt_count'])
                        
                print(f"  ✓ Обработано чеков: {len(all_receipts)}")
            
            # Обрабатываем товарные позиции
            if city_code in cartitems_data and cartitems_data[city_code]:
                all_cartitems = pd.concat(cartitems_data[city_code], ignore_index=True)
                
                if city_code in receipts_data and receipts_data[city_code]:
                    all_receipts = pd.concat(receipts_data[city_code], ignore_index=True)
                    
                    # Связываем товарные позиции с чеками
                    cartitems_with_shops = all_cartitems.merge(
                        all_receipts[['id', 'shop_id']], 
                        left_on='receipt_id', 
                        right_on='id', 
                        how='inner'
                    )
                    
                    if not cartitems_with_shops.empty:
                        # Группируем по магазинам
                        sales_summary = cartitems_with_shops.groupby('shop_id').agg({
                            'total_price': 'sum',
                            'qty': 'sum'
                        }).reset_index()
                        
                        for _, row in sales_summary.iterrows():
                            shop_id = row['shop_id']
                            if shop_id in city_metrics:
                                city_metrics[shop_id]['revenue'] = float(row['total_price'])
                                city_metrics[shop_id]['quantity'] = float(row['qty'])
                                
                        print(f"  ✓ Обработано товарных позиций: {len(all_cartitems)}")
                        print(f"  ✓ Связано с чеками: {len(cartitems_with_shops)}")
            
            period_metrics[city_code] = city_metrics
            
            # Статистика по городу
            total_receipts = sum([m['receipts'] for m in city_metrics.values()])
            total_revenue = sum([m['revenue'] for m in city_metrics.values()])
            total_quantity = sum([m['quantity'] for m in city_metrics.values()])
            
            print(f"  📊 Итого по городу: {total_receipts} чеков, оборот: {total_revenue:.2f}, товаров: {total_quantity:.1f}")
            
        return period_metrics
        
    def create_comparison_report(self, shops_data, period1_metrics, period2_metrics, period1_name, period2_name):
        """Создает сравнительный отчет"""
        print(f"\n📈 СОЗДАНИЕ СРАВНИТЕЛЬНОГО ОТЧЕТА")
        print("-" * 50)
        
        report_data = []
        
        for city_code, city_name in self.cities.items():
            if city_code not in shops_data:
                continue
                
            shops_df = shops_data[city_code]
            
            for _, shop in shops_df.iterrows():
                shop_id = shop['id']
                shop_name = shop['name']
                
                # Метрики первого периода
                p1_metrics = period1_metrics.get(city_code, {}).get(shop_id, {
                    'receipts': 0, 'revenue': 0.0, 'quantity': 0.0
                })
                
                # Метрики второго периода
                p2_metrics = period2_metrics.get(city_code, {}).get(shop_id, {
                    'receipts': 0, 'revenue': 0.0, 'quantity': 0.0
                })
                
                # Абсолютные изменения
                change_receipts = p2_metrics['receipts'] - p1_metrics['receipts']
                change_revenue = p2_metrics['revenue'] - p1_metrics['revenue']
                change_quantity = p2_metrics['quantity'] - p1_metrics['quantity']
                
                # Процентные изменения
                change_receipts_pct = (change_receipts / p1_metrics['receipts'] * 100) if p1_metrics['receipts'] > 0 else 0
                change_revenue_pct = (change_revenue / p1_metrics['revenue'] * 100) if p1_metrics['revenue'] > 0 else 0
                change_quantity_pct = (change_quantity / p1_metrics['quantity'] * 100) if p1_metrics['quantity'] > 0 else 0
                
                report_data.append({
                    'Город': city_name,
                    'ID_магазина': shop_id,
                    'Название_магазина': shop_name,
                    f'{period1_name}_Чеки': p1_metrics['receipts'],
                    f'{period1_name}_Оборот': p1_metrics['revenue'],
                    f'{period1_name}_Количество': p1_metrics['quantity'],
                    f'{period2_name}_Чеки': p2_metrics['receipts'],
                    f'{period2_name}_Оборот': p2_metrics['revenue'],
                    f'{period2_name}_Количество': p2_metrics['quantity'],
                    'Изм_Чеки_абс': change_receipts,
                    'Изм_Оборот_абс': change_revenue,
                    'Изм_Количество_абс': change_quantity,
                    'Изм_Чеки_%': change_receipts_pct,
                    'Изм_Оборот_%': change_revenue_pct,
                    'Изм_Количество_%': change_quantity_pct
                })
                
        return report_data
        
    def save_comparison_excel_report(self, report_data, period1_dates, period2_dates, period1_name, period2_name):
        """Сохраняет сравнительный отчет в Excel"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Создаем DataFrame
        df = pd.DataFrame(report_data)
        
        # Сортируем по городу и изменению оборота
        df = df.sort_values(['Город', 'Изм_Оборот_абс'], ascending=[True, False])
        
        # Форматируем числовые поля
        numeric_cols = [col for col in df.columns if 'Оборот' in col or 'Количество' in col]
        for col in numeric_cols:
            if '%' not in col:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0).round(2)
        
        # Форматируем процентные изменения
        pct_cols = [col for col in df.columns if '%' in col]
        for col in pct_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0).round(1)
        
        # Сохраняем в Excel
        excel_file = self.reports_dir / f"comparison_report_{period1_dates}_{period2_dates}_{timestamp}.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Сравнительный отчет', index=False)
            
            # Получаем объект рабочего листа для форматирования
            worksheet = writer.sheets['Сравнительный отчет']
            
            # Форматируем числовые колонки
            from openpyxl.styles import PatternFill
            
            # Выделяем положительные изменения зеленым, отрицательные красным
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"\n📄 СРАВНИТЕЛЬНЫЙ ОТЧЕТ СОХРАНЕН:")
        print(f"  Excel: {excel_file.absolute()}")
        
        return df
        
    def print_comparison_summary(self, df, period1_name, period2_name):
        """Выводит краткую сводку сравнительного отчета"""
        print(f"\n📈 СВОДКА СРАВНИТЕЛЬНОГО ОТЧЕТА:")
        print("-" * 60)
        
        # Общие метрики
        p1_receipts = df[f'{period1_name}_Чеки'].sum()
        p1_revenue = df[f'{period1_name}_Оборот'].sum()
        p1_quantity = df[f'{period1_name}_Количество'].sum()
        
        p2_receipts = df[f'{period2_name}_Чеки'].sum()
        p2_revenue = df[f'{period2_name}_Оборот'].sum()
        p2_quantity = df[f'{period2_name}_Количество'].sum()
        
        print(f"📊 {period1_name.upper()}:")
        print(f"  🧾 Чеков: {int(p1_receipts):,}")
        print(f"  💰 Оборот: {float(p1_revenue):,.2f}")
        print(f"  📦 Товаров: {float(p1_quantity):,.1f}")
        
        print(f"\n📊 {period2_name.upper()}:")
        print(f"  🧾 Чеков: {int(p2_receipts):,}")
        print(f"  💰 Оборот: {float(p2_revenue):,.2f}")
        print(f"  📦 Товаров: {float(p2_quantity):,.1f}")
        
        # Изменения
        change_receipts = p2_receipts - p1_receipts
        change_revenue = p2_revenue - p1_revenue
        change_quantity = p2_quantity - p1_quantity
        
        change_receipts_pct = (change_receipts / p1_receipts * 100) if p1_receipts > 0 else 0
        change_revenue_pct = (change_revenue / p1_revenue * 100) if p1_revenue > 0 else 0
        change_quantity_pct = (change_quantity / p1_quantity * 100) if p1_quantity > 0 else 0
        
        print(f"\n📈 ИЗМЕНЕНИЯ:")
        print(f"  🧾 Чеки: {int(change_receipts):+,} ({change_receipts_pct:+.1f}%)")
        print(f"  💰 Оборот: {float(change_revenue):+,.2f} ({change_revenue_pct:+.1f}%)")
        print(f"  📦 Товары: {float(change_quantity):+,.1f} ({change_quantity_pct:+.1f}%)")
        
        # ТОП растущих магазинов
        print(f"\n🚀 ТОП-5 РАСТУЩИХ МАГАЗИНОВ ПО ОБОРОТУ:")
        top_growth = df.nlargest(5, 'Изм_Оборот_абс')
        for i, (_, row) in enumerate(top_growth.iterrows(), 1):
            if row['Изм_Оборот_абс'] > 0:
                print(f"  {i}. {row['Название_магазина']} ({row['Город']}) +{float(row['Изм_Оборот_абс']):,.2f} ({float(row['Изм_Оборот_%']):+.1f}%)")
        
        # ТОП падающих магазинов
        print(f"\n📉 ТОП-5 ПАДАЮЩИХ МАГАЗИНОВ ПО ОБОРОТУ:")
        top_decline = df.nsmallest(5, 'Изм_Оборот_абс')
        for i, (_, row) in enumerate(top_decline.iterrows(), 1):
            if row['Изм_Оборот_абс'] < 0:
                print(f"  {i}. {row['Название_магазина']} ({row['Город']}) {float(row['Изм_Оборот_абс']):,.2f} ({float(row['Изм_Оборот_%']):+.1f}%)")
        
        # По городам
        print(f"\n🏙️  ИЗМЕНЕНИЯ ПО ГОРОДАМ:")
        city_summary = df.groupby('Город').agg({
            f'{period1_name}_Оборот': 'sum',
            f'{period2_name}_Оборот': 'sum',
            'Изм_Оборот_абс': 'sum'
        }).round(2)
        
        for city in city_summary.index:
            p1_rev = float(city_summary.loc[city, f'{period1_name}_Оборот'])
            p2_rev = float(city_summary.loc[city, f'{period2_name}_Оборот'])
            change = float(city_summary.loc[city, 'Изм_Оборот_абс'])
            change_pct = (change / p1_rev * 100) if p1_rev > 0 else 0
            print(f"  {city}: {p1_rev:,.2f} → {p2_rev:,.2f} ({change:+,.2f}, {change_pct:+.1f}%)")
            
    def run_comparison_report_generation(self):
        """Запускает генерацию сравнительного отчета"""
        print("=" * 80)
        print("🚀 ГЕНЕРАЦИЯ СРАВНИТЕЛЬНОГО ОТЧЕТА ДВУХ ПЕРИОДОВ")
        print("=" * 80)
        
        # Параметры периодов
        period1_start = "2025-06-02"
        period1_end = "2025-06-08"
        period1_name = "Период1"
        
        period2_start = "2025-06-09"
        period2_end = "2025-06-15"
        period2_name = "Период2"
        
        print(f"📅 {period1_name}: {period1_start} - {period1_end}")
        print(f"📅 {period2_name}: {period2_start} - {period2_end}")
        print(f"🏙️  Города: {', '.join(self.cities.values())}")
        print(f"📊 Сравнение: Количество чеков, Оборот, Количество товаров")
        
        # Спрашиваем об очистке кеша
        if not self.ask_clear_cache():
            return False
            
        # Подключаемся к FTP
        ftp = self.connect_ftp()
        if not ftp:
            print("❌ Не удалось подключиться к FTP-серверу")
            return False
            
        try:
            # Загружаем справочники магазинов
            shops_data = self.download_shop_files(ftp)
            
            # Загружаем данные первого периода
            period1_receipts, period1_cartitems = self.download_period_data(
                ftp, period1_start, period1_end, period1_name
            )
            
            # Загружаем данные второго периода
            period2_receipts, period2_cartitems = self.download_period_data(
                ftp, period2_start, period2_end, period2_name
            )
            
            # Закрываем соединение
            ftp.quit()
            print("\n✓ Соединение с FTP закрыто")
            
            # Рассчитываем метрики для каждого периода
            period1_metrics = self.calculate_period_metrics(
                shops_data, period1_receipts, period1_cartitems, period1_name
            )
            
            period2_metrics = self.calculate_period_metrics(
                shops_data, period2_receipts, period2_cartitems, period2_name
            )
            
            # Создаем сравнительный отчет
            report_data = self.create_comparison_report(
                shops_data, period1_metrics, period2_metrics, period1_name, period2_name
            )
            
            # Сохраняем отчет
            df = self.save_comparison_excel_report(
                report_data, 
                f"{period1_start}_{period1_end}",
                f"{period2_start}_{period2_end}",
                period1_name, 
                period2_name
            )
            
            # Выводим сводку
            self.print_comparison_summary(df, period1_name, period2_name)
            
            print(f"\n✅ СРАВНИТЕЛЬНЫЙ ОТЧЕТ УСПЕШНО СГЕНЕРИРОВАН!")
            print(f"📁 Папка отчетов: {self.reports_dir.absolute()}")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка в процессе генерации отчета: {e}")
            try:
                ftp.quit()
            except:
                pass
            return False

def main():
    """Главная функция программы"""
    print("FTP Analyzer v4.0 - Сравнительный анализатор продаж")
    print("Автор: Система аналитики данных\n")
    
    try:
        analyzer = FTPAnalyzer()
        
        # Проверяем наличие необходимых библиотек
        try:
            import pandas as pd
            import openpyxl
        except ImportError as e:
            print("❌ Отсутствуют необходимые библиотеки!")
            print("Установите их командами:")
            print("pip install pandas openpyxl")
            return
        
        # Сразу запускаем генерацию сравнительного отчета
        analyzer.run_comparison_report_generation()
                
    except KeyboardInterrupt:
        print("\n\n⏹️  Программа прервана пользователем")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")

if __name__ == "__main__":
    main()